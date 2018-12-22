# -*- coding:utf8 -*-

'''
运行于  Python3 .
对维护的图书进行规范化命名。
对于没有 ID 的，添加 ID.
'''

import os
import sys
import random
import re
import shutil
from openpyxl import load_workbook

XLSX_META = './database/books/xx_books.xlsx'


def get_uu4d():
    sel_arr = ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9']
    slice = random.sample(sel_arr, 4)
    return ('k' + ''.join(slice))


def get_sig_arr():
    '''
    得到所有的数据集编码列表
    '''
    sig_arr = []
    # pattern='dataset.+_tt$'

    wb = load_workbook(XLSX_META)
    sheets = wb.sheetnames
    for sheet in sheets:
        ws = wb[sheet]
        rows = ws.max_row
        for i in range(2, rows + 1):
            sig = ws.cell(row=i, column=1).value
            if sig:
                pass
            else:
                continue
            sig_arr.append(sig)

    print(sig_arr)
    return sig_arr


def do_for_dt():
    current_sig_arr = get_sig_arr()

    wb = load_workbook(XLSX_META)
    sheets = wb.sheetnames
    for sheet in sheets:
        ws = wb[sheet]
        rows = ws.max_row
        for i in range(2, rows + 1):
            sig = ws.cell(row=i, column=1).value
            if sig:
                continue
            else:
                new_sig = get_uu4d()
                while new_sig in current_sig_arr:
                    new_sig = get_uu4d()

                ws.cell(row=i, column=1).value = new_sig

            current_sig_arr.append(new_sig)
    wb.save(XLSX_META)

    print(current_sig_arr)


if __name__ == '__main__':
    do_for_dt()
