import os

from openpyxl import load_workbook

from urllib.error import HTTPError
import requests
import time
import random


def import_meta():
    wb = load_workbook('./xx_books.xlsx')
    sheets = wb.sheetnames
    for sheet in sheets:
        catid = sheet.split('_')[0]
        ws = wb[sheet]
        rows = ws.max_row
        for i in range(2, rows + 1):
            sig = ws.cell(row=i, column=1).value
            url = ws.cell(row=i, column=4).value
            print(sig)

            outws = os.path.join('.', sig)
            if os.path.exists(outws):
                pass
            else:
                os.mkdir(outws)
            if url:
                getNews(outws, url)


def getNews(outws, url):
    headers = {'User-Agent': 'Mozilla/6.0 (Windows NT 8.1) Chrome/58.0.2305.2',
               'Host': 'www.amazon.cn',
               'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
               'Accept - Language': 'zh - CN, zh;q = 0.5',
               'Accept - Encoding': 'gzip, deflate, br',
               # 'Referer': 'https://www.amazon.cn/dp/B06XKQ3SMQ/ref=sr_1_170?ie=UTF8&qid=1545207798&sr=8-170&keywords=%E5%8E%86%E5%8F%B2%E5%9C%B0%E5%9B%BE',
               'Cookie': 'csm-hit=tb:REJCv96GZGF0WBHRH5RQ+s-Cv96GRZFEJH5RQ0WBHGR|1545377404948&t:1545377404948&adb:adblk_no',
               'Connection': 'keep-alive',
               'Upgrade-Insecure-Requests': '1'
               }

    # 把该url添加进visited()
    try:

        # 只获取一次
        all_text_file = os.path.join(os.path.join(outws, 'all_content.html'))

        if os.path.exists(all_text_file):
            return True
        else:
            html = requests.get(url, headers=headers, )
            print('=' * 40)
            print(html)
            print('=' * 40)
            time.sleep(20 + random.randint(1, 30))

            html = html.text
            with open(all_text_file, 'w') as fo:
                fo.write(html)




    except HTTPError as e:
        print(e)


if __name__ == '__main__':
    import_meta()
