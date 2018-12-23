import os

from openpyxl import load_workbook
from bs4 import BeautifulSoup

from openpyxl import Workbook


def import_meta():
    wb = load_workbook('./xx_books.xlsx')
    sheets = wb.sheetnames
    for sheet in sheets:
        # catid = sheet.split('_')[0]
        ws = wb[sheet]
        rows = ws.max_row
        for i in range(2, rows + 1):
            sig = ws.cell(row=i, column=1).value
            print(sig)

            getNews(sig)


def get_text(obj):
    return obj.text if obj else ''


def clean_str(obj):
    tt = obj.split('\n')
    tt = [x.strip() for x in tt]
    return '\n'.join(tt)


def getNews(sig):
    print('=' * 40)
    outws = os.path.join('.', sig)
    all_text_file = os.path.join(os.path.join(outws, 'all_content.html'))

    if os.path.exists(all_text_file):
        html = open(all_text_file).read()
    else:
        return None

    soup = BeautifulSoup(html, 'html.parser')

    title_s = soup.find('span', {'id': 'productTitle'})
    if title_s:
        pass
    else:
        title_s = soup.find('span', {'id': 'ebooksProductTitle'})
    paper = title_s.text
    print(paper)

    author = get_text(soup.find('span', {'class': 'author'}))

    introduction = soup.find('div', {'id': 'bookDescription_feature_div'})

    if introduction:
        introduction = introduction.noscript
    else:
        introduction = ''
    introduction = get_text(introduction)

    detail_desc = get_text(
        soup.find('div', {'id': 's_contents'})
    )

    detail_desc = clean_str(detail_desc)

    # print(detail_desc)

    wb = Workbook()
    ws = wb.active
    ws.title = "meta"
    # ws = wb['meta']
    ws.cell(row=1, column=1).value = 'title'
    ws.cell(row=1, column=2).value = paper

    ws.cell(row=2, column=1).value = 'author'
    ws.cell(row=2, column=2).value = author

    details = soup.find('div', {'id': 'detail_bullets_id'}).find_all('li')

    idx = 6
    de_str = '### 详情\n'
    for de in details:
        # print('-' * 20)
        qq = de.find('b')
        cccc = ['亚马逊热销商品排名', '用户评分']
        tm = False
        for tt in cccc:
            if tt in qq.text:
                tm = True
        if tm:
            break

        print(qq.text.strip(), ''.join(de.text.split()).strip()[len(qq.text.strip()):])
        ws.cell(row=idx, column=1).value = qq.text.strip()
        ws.cell(row=idx, column=2).value = ''.join(de.text.split()).strip()[len(qq.text.strip()):]
        de_str = de_str + '- *{0}* {1}\n'.format(qq.text.strip(),
                                                 ''.join(de.text.split()).strip()[len(qq.text.strip()):])
        idx = idx + 1



    wb.save(os.path.join(outws, 'meta.xlsx'))

    with open(os.path.join(outws, 'introduction.md'), 'w') as fo:
        fo.write('### 介绍\n')
        fo.write(introduction)
        fo.write('\n')
        fo.write(detail_desc)
        fo.write('\n')
        fo.write(de_str)


if __name__ == '__main__':
    import_meta()
