# -*- coding: utf-8

'''
Import books metadata which ID extracted from XLSX.
'''

import sys
import os

import pathlib
from openpyxl import load_workbook

from torcms.model.category_model import MCategory
from torcms.model.post2catalog_model import MPost2Catalog
from torcms.model.post_model import MPost


# import xlrd  # 删除


def update_category(uid, postdata, kwargs):
    '''
    Update the category of the post.
    '''
    catid = kwargs['catid'] if ('catid' in kwargs and MCategory.get_by_uid(kwargs['catid'])) else None

    post_data = postdata

    current_infos = MPost2Catalog.query_by_entity_uid(uid, kind='').objects()

    new_category_arr = []
    # Used to update post2category, to keep order.
    def_cate_arr = ['gcat{0}'.format(x) for x in range(10)]

    # for old page.
    def_cate_arr.append('def_cat_uid')

    # Used to update post extinfo.
    cat_dic = {}
    for key in def_cate_arr:
        if key not in post_data:
            continue
        if post_data[key] == '' or post_data[key] == '0':
            continue
        # 有可能选重复了。保留前面的
        if post_data[key] in new_category_arr:
            continue

        new_category_arr.append(post_data[key] + ' ' * (4 - len(post_data[key])))
        cat_dic[key] = post_data[key] + ' ' * (4 - len(post_data[key]))

    if catid:
        def_cat_id = catid
    elif new_category_arr:
        def_cat_id = new_category_arr[0]
    else:
        def_cat_id = None

    if def_cat_id:
        cat_dic['def_cat_uid'] = def_cat_id
        cat_dic['def_cat_pid'] = MCategory.get_by_uid(def_cat_id).pid

    print('=' * 40)
    print(uid)
    print(cat_dic)
    MPost.update_jsonb(uid, cat_dic)

    for index, catid in enumerate(new_category_arr):
        MPost2Catalog.add_record(uid, catid, index)

    # Delete the old category if not in post requests.
    for cur_info in current_infos:
        if cur_info.tag_id not in new_category_arr:
            MPost2Catalog.remove_relation(uid, cur_info.tag_id)


def chuli_meta(sig, metafile):
    print(metafile)
    wb = load_workbook(str(metafile))
    sheet = wb[wb.sheetnames[0]]
    meta_dic = {}
    for row in sheet.iter_rows():
        the_key = row[0].value
        the_val = row[1].value if row[1].value else ''
        meta_dic[the_key] = the_val
        # print()
    meta_dic['identifier'] = sig
    print(meta_dic)
    # mrec.add_or_update(meta_dic)
    return meta_dic


def get_meta(catid, sig, cell_cars):
    uid = sig[2:]
    meta_base = './static/xx_books_meta'
    if os.path.exists(meta_base):
        pass
    else:
        return False

    for wroot, wdirs, wfiles in os.walk(meta_base):
        for wdir in wdirs:
            # if wdir.lower().endswith(sig):
            ds_base = pathlib.Path(os.path.join(wroot, wdir))

            pp_data = {'logo': '', 'kind': 'k'}
            for uu in ds_base.iterdir():
                if uu.name.endswith('.xlsx'):
                    meta_dic = chuli_meta('u' + sig[2:], uu)

                    pp_data['title'] = cell_cars[1]
                    pp_data['cnt_md'] = cell_cars[1]

                    # pp_data['title'] = meta_dic['title']
                    # pp_data['cnt_md'] = meta_dic['anytext']
                    pp_data['user_name'] = 'admin'
                    pp_data['def_cat_uid'] = catid
                    pp_data['def_cat_pid'] = catid[:2] + '00'

                    pp_data['extinfo'] = {
                        'zlink': cell_cars[3]
                    }

                    print(uu.name)

                    # print(xlsx.absolute())
                    # MPost.update_misc(post_uid, def_tab_path = xlsx.absolute())

                elif uu.name.startswith('thumbnail_'):
                    pp_data['logo'] = os.path.join(wroot, wdir, uu.name).strip('.')

            kwargsa = {
                'def_cat_uid': catid,
                'cat_id': catid,
            }
            MPost.add_or_update('k' + uid, pp_data)
            update_category('k' + uid, pp_data, kwargsa)
            MPost2Catalog.add_record('k' + uid, catid)


def import_meta():
    inws = pathlib.Path('./database/books')
    print(inws)
    # for y in os.listdir(inws):
    #     print(y)
    for cat_path in inws.iterdir():
        print('x' * 40)
        print(cat_path.name)
        if cat_path.name.lower().endswith('.xlsx'):
            print('////')
            pass
        else:
            continue
        wb = load_workbook(str(cat_path))
        sheets = wb.sheetnames
        for sheet in sheets:
            catid = sheet.split('_')[0]
            ws = wb[sheet]
            rows = ws.max_row
            cols = ws.max_column
            for i in range(2, rows + 1):
                sig = ws.cell(row=i, column=1).value
                if sig:
                    pass
                else:
                    continue
                cell_vars = [ws.cell(row=i, column=tt).value for tt in range(1, cols + 1)]
                print(sig)
                if sig:
                    pass
                    get_meta(catid, sig, cell_vars)


if __name__ == '__main__':
    import_meta()
